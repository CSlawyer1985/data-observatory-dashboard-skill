#!/usr/bin/env python3
"""Profile CSV/JSON/XLSX data for data observatory dashboards."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


LABEL_HINTS = ("name", "title", "label", "entity", "company", "occupation", "case", "项目", "名称", "标题", "职业", "公司", "案件")
GROUP_HINTS = ("category", "group", "type", "industry", "region", "department", "jurisdiction", "类别", "分类", "类型", "行业", "地区", "部门")
SIZE_HINTS = ("count", "amount", "value", "revenue", "fee", "income", "assets", "employment", "population", "cases", "total", "数量", "金额", "规模", "人数", "收入", "收费", "创收", "总计", "案件", "件数")
SCORE_HINTS = ("risk", "score", "exposure", "sentiment", "growth", "probability", "intensity", "efficiency", "ratio", "per_", "average", "风险", "评分", "得分", "暴露", "增长", "强度", "效率", "占比", "人均", "平均")
TEXT_HINTS = ("summary", "rationale", "description", "note", "remark", "摘要", "理由", "说明", "备注")
SOURCE_HINTS = ("source", "url", "link", "citation", "来源", "链接")


def load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        obj = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(obj, list):
            return [r for r in obj if isinstance(r, dict)]
        if isinstance(obj, dict):
            for key in ("data", "rows", "records", "items"):
                if isinstance(obj.get(key), list):
                    return [r for r in obj[key] if isinstance(r, dict)]
        raise SystemExit("JSON must be an array of objects or contain data/rows/records/items.")
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    if suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("XLSX support requires openpyxl.") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(rows[0])]
        return [dict(zip(headers, row)) for row in rows[1:]]
    raise SystemExit(f"Unsupported file type: {suffix}")


def is_missing(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def parse_number(v: Any) -> float | None:
    if is_missing(v):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if math.isfinite(float(v)):
            return float(v)
        return None
    s = str(v).strip().replace(",", "")
    s = re.sub(r"[%￥$元万亿人件条个]+$", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def looks_date(v: Any) -> bool:
    if is_missing(v):
        return False
    s = str(v).strip()
    return bool(re.match(r"^\d{4}[-/.年]\d{1,2}([-/.\u6708]\d{1,2})?", s))


def hint_score(field: str, hints: tuple[str, ...]) -> int:
    f = field.lower()
    return sum(1 for h in hints if h.lower() in f)


def classify_field(field: str, values: list[Any], total_rows: int, include_samples: bool = False) -> dict[str, Any]:
    non_missing = [v for v in values if not is_missing(v)]
    missing = total_rows - len(non_missing)
    unique = len({str(v) for v in non_missing})
    nums = [parse_number(v) for v in non_missing]
    nums = [v for v in nums if v is not None]
    numeric_ratio = len(nums) / len(non_missing) if non_missing else 0
    date_ratio = sum(1 for v in non_missing if looks_date(v)) / len(non_missing) if non_missing else 0
    avg_len = sum(len(str(v)) for v in non_missing) / len(non_missing) if non_missing else 0

    if numeric_ratio >= 0.75:
        kind = "numeric"
    elif date_ratio >= 0.6:
        kind = "date"
    elif unique <= max(20, total_rows * 0.25):
        kind = "categorical"
    elif avg_len >= 40:
        kind = "text"
    else:
        kind = "identifier"

    out: dict[str, Any] = {
        "field": field,
        "type": kind,
        "missing": missing,
        "missing_pct": round(missing / total_rows * 100, 2) if total_rows else 0,
        "unique": unique,
    }
    if include_samples:
        out["sample_values"] = [str(v) for v in non_missing[:5]]
    if nums:
        nums_sorted = sorted(nums)
        out["numeric"] = {
            "min": nums_sorted[0],
            "max": nums_sorted[-1],
            "mean": sum(nums_sorted) / len(nums_sorted),
            "positive_count": sum(1 for n in nums_sorted if n > 0),
        }
    if kind == "categorical":
        out["top_values"] = Counter(str(v) for v in non_missing).most_common(10)
    return out


def recommend(fields: list[dict[str, Any]]) -> dict[str, Any]:
    def candidates(hints: tuple[str, ...], kinds: tuple[str, ...], positive_required: bool = False) -> list[str]:
        scored: list[tuple[int, str]] = []
        for f in fields:
            if f["type"] not in kinds:
                continue
            if positive_required and f.get("numeric", {}).get("positive_count", 0) == 0:
                continue
            score = hint_score(f["field"], hints) * 10
            score += max(0, 5 - f.get("missing_pct", 0) / 20)
            if f["type"] == "categorical" and f.get("unique", 999) <= 30:
                score += 2
            scored.append((int(score), f["field"]))
        return [name for _, name in sorted(scored, reverse=True)]

    label = candidates(LABEL_HINTS, ("identifier", "categorical", "text"))
    group = candidates(GROUP_HINTS, ("categorical",))
    size = candidates(SIZE_HINTS, ("numeric",), positive_required=True)
    score = candidates(SCORE_HINTS, ("numeric",), positive_required=False)
    text = candidates(TEXT_HINTS, ("text", "identifier", "categorical"))
    source = candidates(SOURCE_HINTS, ("identifier", "text", "categorical"))

    return {
        "label_candidates": label[:8],
        "group_candidates": group[:8],
        "size_candidates": size[:8],
        "score_color_candidates": score[:8],
        "text_detail_candidates": text[:8],
        "source_candidates": source[:8],
        "suggested_visuals": suggested_visuals(fields, group, size, score),
    }


def suggested_visuals(fields: list[dict[str, Any]], groups: list[str], sizes: list[str], scores: list[str]) -> list[str]:
    has_date = any(f["type"] == "date" for f in fields)
    numeric_count = sum(1 for f in fields if f["type"] == "numeric")
    out = []
    if groups and sizes:
        out.append("treemap observatory")
    if numeric_count >= 2:
        out.append("scatter/bubble observatory")
    if has_date:
        out.append("timeline or small-multiple trend dashboard")
    if groups:
        out.append("dense grouped matrix")
    if sizes:
        out.append("ranked bar dashboard")
    return out or ["summary table dashboard"]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--out", type=Path, default=Path("data-profile.json"))
    parser.add_argument("--include-samples", action="store_true", help="Include raw sample values in the profile. Off by default to avoid accidental sensitive-data leakage.")
    args = parser.parse_args()

    rows = load_rows(args.input)
    if not rows:
        raise SystemExit("No rows found.")

    fields = sorted({k for row in rows for k in row.keys()})
    profiled = [classify_field(field, [row.get(field) for row in rows], len(rows), args.include_samples) for field in fields]
    output = {
        "input": args.input.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "row_count": len(rows),
        "field_count": len(fields),
        "fields": profiled,
        "recommendations": recommend(profiled),
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
